import openpyxl
import math
import turtle


def get_data(sheet, row):
    data = list(map(lambda x: float(x.value), list(sheet.rows)[row - 1]))
    #
    return data


def data_in_sheet(al, v, s, u1n, u2n, w1, w2, sp1, sp2, d1, a1, a2, a12, l, d, pc, pi,
                  hc, hi, c, pk, px, uk, i0, u_sheet):
    u_sheet.column_dimensions['B'].width = 20
    u_sheet.column_dimensions['A'].width = 30
    u_sheet['A1'] = f'Вариант: {v}'
    u_sheet['A2'] = 'Исходные данные: '
    u_sheet['A3'] = '1. Обмотка алюминиевая.' if al else 'Обмотка медная.'
    u_sheet['A4'] = f'2. Мощность S = {s} кВА'
    u_sheet['A4'] = '3. Схемы и группа соединений   Y/Y-0'
    u_sheet.merge_cells('A4:B4')
    u_sheet['A5'] = '4. Номинальные напряжения, В:'
    u_sheet['A6'] = 'ВН:'
    u_sheet['A7'] = 'НН:'
    u_sheet['B6'] = f'U1ном = {u1n}'
    u_sheet['B7'] = f'U2ном = {u2n}'
    u_sheet['A8'] = '5. Обмотки:'
    u_sheet['A9'] = 'Число витков:'
    u_sheet['A10'] = u_sheet['A13'] = u_sheet['A17'] = 'ВН:'
    u_sheet['A11'] = u_sheet['A14'] = u_sheet['A18'] = 'НН:'
    u_sheet['B10'] = f'w1 = {w1}'
    u_sheet['B11'] = f'w2 = {w2}'
    u_sheet['A12'] = 'Сечение витков, мм2:'
    u_sheet['B13'] = f'Sпр1 = {sp1} мм2'
    u_sheet['B14'] = f'Sпр2 = {sp2} мм2'
    u_sheet['A15'] = f'Внутренний диаметр обмотки НН D1 = {d1} см'
    u_sheet.merge_cells('A15:B15')
    u_sheet['A16'] = 'Радиальные размеры, см: '
    u_sheet['B17'] = f'а1 = {a1} см'
    u_sheet['B18'] = f'а2 = {a2} см'
    u_sheet['A19'] = f'Канал между обмотками ВН и НН  а12 = {a12} см'
    u_sheet.merge_cells('A19:B19')
    u_sheet['A20'] = f'Высота обмотки, l1=l2,см		l1 = {l} см'
    u_sheet.merge_cells('A20:B20')
    u_sheet['A21'] = '6.	Магнитная система'
    u_sheet['A22'] = f'Диаметр стержня d, см    	d = {d} см'
    u_sheet['A23'] = 'Активное сечение, см2   '
    u_sheet['A24'] = 'Стержень'
    u_sheet['B24'] = f'Пc  = {pc} см2'
    u_sheet['A25'] = 'Ярмо '
    u_sheet['B25'] = f'Пя  = {pi} см2'
    u_sheet['A26'] = 'Высота, см'
    u_sheet['A27'] = 'Стержень'
    u_sheet['B27'] = f'hc  = {hc} см'
    u_sheet['A28'] = 'Ярмо '
    u_sheet['B28'] = f'hя  = {hi} см'
    u_sheet['A29'] = f'Расстояние между осями С = {c} см'
    u_sheet['A29'] = '7.	Контрольные данные'
    u_sheet['A30'] = f'Pк = {pk} Вт'
    u_sheet['A31'] = f'uк = {uk}%'
    u_sheet['B30'] = f'P_х={px}0 Вт'
    u_sheet['B31'] = f'i0 = {i0}%'


def get_accessory_table(sheet):
    accessory = [list(map(lambda x: float(x.value.replace(',', '.')), list(sheet.columns)[column])) for column in
                 range(0, 4)]
    return accessory


def nearest_value(items, value):
    return min(items, key=lambda x: abs(value - abs(x)))


def step_1(u1n, u2n, s, w1, w2, c_sheet):
    c_sheet['A1'] = '1'
    u1fn = u1n / 3 ** (1 / 2)
    u2fn = u2n / 3 ** (1 / 2)
    c_sheet['A1'] = 'Фазные значения номинального напряжения (при соединении Y/Y)'
    c_sheet.merge_cells('A1:G1')
    c_sheet['A2'] = f'U_1фном=U_1ном/√3={u1n}/√3='
    c_sheet['I2'] = f'{round(u1fn, 2)} B'
    c_sheet.merge_cells('A2:D2')
    c_sheet['A3'] = f'U_2фном=U_2ном/√3={u2n}/√3='
    c_sheet['I3'] = f'{round(u2fn, 2)} B'
    c_sheet.merge_cells('A3:D3')

    i1n = s * 1000 / (3 ** (1 / 2) * u1n)
    i2n = s * 1000 / (3 ** (1 / 2) * u2n)
    c_sheet['A5'] = 'Номинальные и фазные значения токов'
    c_sheet.merge_cells('A5:G5')
    c_sheet['A6'] = f'I_1ном=I_1ф=S/(√3∙U_1 )={s * 1e3}/(√3∙{u1n})='
    c_sheet['I6'] = f'{round(i1n, 2)} A'
    c_sheet.merge_cells('A6:E6')
    c_sheet['A7'] = f'I_1ном=I_1ф=S/(√3∙U_2 )={s * 1e3}/(√3∙{u2n})='
    c_sheet['I7'] = f'{round(i2n, 2)} A'
    c_sheet.merge_cells('A7:E7')

    k = w1 / w2
    c_sheet['A9'] = 'Коэффициент трансформации'
    c_sheet.merge_cells('A5:G5')
    c_sheet['A10'] = f'k=w_1/w_2 ={w1}/{w2}'
    c_sheet['I10'] = f'{round(k, 2)}'
    c_sheet.merge_cells('A10:D10')
    return u1fn, u2fn, i1n, i2n, k


s = turtle.getscreen()
t = turtle.Turtle()
t.speed(0)
t.hideturtle()
t.screen.screensize(1000, 1000)


def step_2(data, v):
    t.clear()
    m = 2
    width = 2 * data[-1] + 2 * (data[-2] - data[-4] / 2)
    for ms in (1, 2, 2.5, 4, 5, 10, 15, 20, 25, 40, 50):
        if width / ms <= 15:
            break
    d1, a1, a2, a12, l, d, hc, hi, c = list(map(lambda x: m * 10 * x / ms, data))
    t.pencolor('blue')
    cletk = 35
    kwadr = 5 * m
    t.penup()

    t.goto(-kwadr - c - (hi - d / 2), kwadr + hc / 2 + hi)
    t.pendown()
    for _ in range(2):
        for i in (cletk // 2) * (1, -1):
            t.rt(i * 90)
            t.forward(cletk * kwadr)
            t.rt(-i * 90)
            t.forward(kwadr)
        t.rt(90)
    t.penup()
    t.home()

    t.pencolor('black')
    t.goto(-c - (hi - d / 2), hc / 2 + hi)
    t.pendown()
    rect(2 * c + 2 * (hi - d / 2), hc + 2 * hi)  # outer rect

    t.penup()
    tp = t.pos()
    t.goto(tp[0] + hi / 2 + d / 2, tp[1] - hi)
    t.pendown()

    coil(hc, a1, a12, a2, (d1 - d) / 2, l)

    t.penup()
    t.goto(c - d / 2, - hc / 2)
    t.pendown()
    coil(-hc, -a1, -a12, -a2, -(d1 - d) / 2, -l)
    turtle.getscreen().getcanvas().postscript(file=f'var{int(v)}_M1:{ms}.eps')


def rect(width, lenght):
    for forw in 2 * (width, lenght):
        t.fd(forw)
        t.rt(90)


def coil(c_l, a1, a12, a2, a_, l):
    c_w = 2 * (a1 + a12 + a2)
    rect(c_w, c_l)
    t.penup()
    tp = t.pos()
    t.goto(tp[0] + a1 + a12 + a2 - a_, tp[1] - (c_l - l) / 2)
    t.pendown()
    rect(a1 + a12 + a2, l)
    t.fd(a2)
    rect(a12, l)

    # Scetch.
    pass


def step_3(r_sheet, s, i1n, u1fn, w1, pc, pi, hc, c, d, f, kd, accessory, c_sheet, r=7650):
    # Idling experience.
    li = 2 * c + d
    gc = round(3 * hc * pc * r * 10 ** -6, 2)
    gi = round(2 * li * pi * r * 10 ** -6, 2)
    c_sheet['A13'] = 'Массы стали стержней и ярм:'
    c_sheet.merge_cells('A13:E13')
    c_sheet['A14'] = f'G_с=3h_с∙ρ_ст∙П_с=3∙{hc}∙{pc}∙{r}∙10^(-6)='
    c_sheet.merge_cells('A14:E14')
    c_sheet['I14'] = f'{round(gc, 2)} кг'
    c_sheet['A15'] = f'l_я=2∙c+d=2∙{c}+{d}='
    c_sheet.merge_cells('A15:E15')
    c_sheet['I15'] = f'{round(li, 2)} cm'
    c_sheet['A16'] = f'G_я=2∙l_я∙ρ_ст∙П_я=2∙{li}∙{pi}∙{r}∙10^(-6)='
    c_sheet.merge_cells('A16:E16')
    c_sheet['I16'] = f'{round(gi, 2)} кг'
    c_sheet['A17'] = 'Плотность холоднокатаной стали ρ_ст=7650 кг/м3'
    c_sheet.merge_cells('A17:E17')

    r_sheet['A1'] = 'n'
    r_sheet['B1'] = 'U1f, V'
    r_sheet['C1'] = 'Px, W'
    r_sheet['D1'] = 'Qx, WAp'
    r_sheet['E1'] = 'I0a, A'
    r_sheet['F1'] = 'I0p, A'
    r_sheet['G1'] = 'I0, A'
    r_sheet['H1'] = 'cosp0'

    row_t2 = 2
    for n in (0.5, 0.7, 0.8, 0.9, 1, 1.1):
        tabl_data = step_3_tabl(n, u1fn, i1n, w1, pc, pi, f, kd, accessory, gc, gi, s, c_sheet)
        if n == 1:
            r_sheet['K1'] = 'Результаты вычислений'
            r_sheet['K2'] = f'{tabl_data[0]} W'
            r_sheet['K3'] = f'{tabl_data[-5]} %'
            bc = tabl_data[-4]
            px = tabl_data[0]
            r0 = tabl_data[-2]
            x0 = tabl_data[-1]
        r_sheet[f'A{row_t2}'] = n
        r_sheet[f'B{row_t2}'] = round(n * u1fn, 2)
        r_sheet[f'C{row_t2}'] = tabl_data[0]
        r_sheet[f'D{row_t2}'] = tabl_data[1]
        r_sheet[f'E{row_t2}'] = tabl_data[2]
        r_sheet[f'F{row_t2}'] = tabl_data[3]
        r_sheet[f'G{row_t2}'] = tabl_data[4]
        r_sheet[f'H{row_t2}'] = tabl_data[5]
        row_t2 += 1
    bc40 = round(u1fn * 10 ** 4 / (4.44 * 40 * w1 * pc), 3)
    bc60 = round(u1fn * 10 ** 4 / (4.44 * 60 * w1 * pc), 3)
    px40 = round(px * (bc40 / bc) ** 2 * (40 / 50) ** 1.5, 2)
    px60 = round(px * (bc60 / bc) ** 2 * (60 / 50) ** 1.5, 2)
    c_sheet['A51'] = 'P_х=P_х50∙(B/B_50 )^2∙(f/f_50 )^1.5'
    c_sheet.merge_cells('A51:G51')
    c_sheet['A52'] = 'где Рx50, В50  и  f50 — данные при частоте 50 Гц.'
    c_sheet.merge_cells('A52:G52')
    c_sheet['A53'] = 'Индукция при частоте 40 и 60 Гц'
    c_sheet.merge_cells('A53:G53')
    c_sheet['A54'] = f'B_с40=(U_1фном∙10^4)/(4.44∙f_40∙w_1∙П_с )=({round(u1fn, 2)}∙10^4)/(4.44∙40∙{w1}∙{pc})='
    c_sheet.merge_cells('A54:G54')
    c_sheet['I54'] = f'{round(bc40, 2)} Тл'
    c_sheet['A55'] = f'B_с60=(U_1фном∙10^4)/(4.44∙f_60∙w_1∙П_с )=({round(u1fn, 2)}∙10^4)/(4.44∙60∙{w1}∙{pc})='
    c_sheet.merge_cells('A55:G55')
    c_sheet['I55'] = f'{round(bc60, 2)} Тл'
    c_sheet['A56'] = 'Потери холостого хода:'
    c_sheet.merge_cells('A56:G56')
    c_sheet['A57'] = f'P_x40={round(px, 2)}∙({round(bc40, 2)}/{round(bc, 2)})^2 (40/50)^1.5='
    c_sheet.merge_cells('A57:G57')
    c_sheet['I57'] = f'{round(px40)} Вт'
    c_sheet['A58'] = f'P_x60={round(px, 2)}∙({round(bc60, 2)}/{round(bc, 2)})^2 (60/50)^1.5='
    c_sheet.merge_cells('A58:G58')
    c_sheet['I58'] = f'{round(px60)} Вт'

    return r0, x0, px


def step_3_tabl(n, u1fn, i1n, w1, pc, pi, f, kd, accessory, gc, gi, s, c_sheet):
    u1fn *= n
    bc = round(u1fn * 10 ** 4 / (4.44 * f * w1 * pc), 3)
    bi = round(u1fn * 10 ** 4 / (4.44 * f * w1 * pi), 3)
    c_ac_ind = accessory[0].index(nearest_value(accessory[0], bc))
    i_ac_ind = accessory[0].index(nearest_value(accessory[0], bi))
    pc_ac = accessory[1][c_ac_ind]
    qc_ac = accessory[2][c_ac_ind]
    qzc_ac = accessory[3][c_ac_ind]
    pi_ac = accessory[1][i_ac_ind]
    qi_ac = accessory[2][i_ac_ind]
    qzi_ac = accessory[3][i_ac_ind]
    px = round(kd * (pc_ac * gc + pi_ac * gi), 2)
    qx = round(qc_ac * gc + qi_ac * gi + qzc_ac * 3 * pc + qzi_ac * 4 * pi, 2)

    i_0a = round(px / (10 * s), 2)
    i_0p = round(qx / (10 * s), 2)
    i_0 = round((i_0a ** 2 + i_0p ** 2) ** (1 / 2), 2)
    i0a = round(i_0a * i1n / 100, 5)
    i0p = round(i_0p * i1n / 100, 5)
    i0 = round(i_0 * i1n / 100, 5)
    cosp0 = round(i0a / i0, 3)
    sinp0 = round(i0p / i0, 3)

    z0 = round(u1fn / i0, 2)
    r0 = round(z0 * cosp0, 2)
    x0 = round((z0 ** 2 - r0 ** 2) ** (1 / 2), 2)
    if n == 1:
        c_sheet['A18'] = 'Индукции в стержнях и ярмах'
        c_sheet.merge_cells('A18:E18')
        c_sheet['A19'] = f'B_с=(U_1фном∙10^4)/(4.44∙f∙w_1∙П_с )=({round(u1fn, 2)}∙10^4)/(4.44∙50∙{w1}∙{pc})='
        c_sheet.merge_cells('A19:G19')
        c_sheet['I19'] = f'{round(bc, 2)} Тл'
        c_sheet['A20'] = f'B_с=(U_1фном∙10^4)/(4.44∙f∙w_1∙П_с )=({round(u1fn, 2)}∙10^4)/(4.44∙50∙{w1}∙{pi})='
        c_sheet.merge_cells('A20:G20')
        c_sheet['I20'] = f'{round(bi, 2)} Тл'
        c_sheet['A21'] = 'Удельные потери в стали стержней и ярм, определенные по табл. 3 '
        c_sheet.merge_cells('A21:G21')
        c_sheet['A22'] = f'                        p_с={pc_ac} Вт/кг          p_я={pi_ac} Вт/кг'
        c_sheet.merge_cells('A22:G22')
        c_sheet['A23'] = 'потери холостого хода Рx:'
        c_sheet.merge_cells('A23:E23')
        c_sheet['A24'] = f'P_х=K_д∙(p_с∙G_с+p_я∙G_я )=1.25∙({pc}∙{round(gc, 2)}+{pi}∙{round(gi, 2)})='
        c_sheet.merge_cells('A24:G24')
        c_sheet['I24'] = f'{round(px, 2)} Вт'
        c_sheet['A25'] = 'где K_д=1.25'

        c_sheet['A26'] = 'По таблице 3 выбираем значения q_с и q_я, q_зс и q_зя'
        c_sheet.merge_cells('A26:G26')
        c_sheet['A27'] = f'          q_с={qc_ac} В·А/кг          q_зс={qzc_ac} В·А/см2'
        c_sheet.merge_cells('A27:G27')
        c_sheet['A28'] = f'          q_i={qi_ac} В·А/кг          q_зi={qzi_ac} В·А/см2'
        c_sheet.merge_cells('A28:G28')
        c_sheet['A29'] = 'Намагничивающая мощность магнитной системы ( реактивная мощность х.х.)'
        c_sheet.merge_cells('A29:G29')
        c_sheet['A30'] = 'Q_х=q_с∙G_с+q_я∙G_я+q_зс∙n_зс∙П_с+q_зя∙n_зя∙П_я'
        c_sheet.merge_cells('A30:G30')
        c_sheet['A31'] = f'Q_х={qc_ac}∙{round(gc, 2)}+{qi_ac}∙{round(gi, 2)}+{qzc_ac}∙3∙{pc}+{qzi_ac}∙4∙{pi}='
        c_sheet.merge_cells('A31:G31')
        c_sheet['I31'] = f'{qx} ВАр'

        c_sheet['A33'] = 'Ток холостого хода:'
        c_sheet.merge_cells('A33:G33')
        c_sheet['A34'] = 'активная составляющая, %'
        c_sheet.merge_cells('A34:G34')
        c_sheet['A35'] = f'i_0а=P_х/(10∙S)={round(px, 2)}/(10∙{round(s, 2)})='
        c_sheet.merge_cells('A35:G35')
        c_sheet['I35'] = f'{round(i_0a, 2)} %.'
        c_sheet['A36'] = 'реактивная составляющая, %'
        c_sheet.merge_cells('A36:G36')
        c_sheet['A37'] = f'i_0р=Q_х/(10∙S)={round(qx, 2)}/(10∙{round(s, 2)})='
        c_sheet.merge_cells('A37:G37')
        c_sheet['I37'] = f'{round(i_0p, 2)} %.'
        c_sheet['A38'] = 'полный ток холостого хода'
        c_sheet.merge_cells('A38:G38')
        c_sheet['A38'] = f'i_0=√(i_0а^2+i_0р^2 )=√({round(i_0a)}^2+{round(i_0p)}^2 )='
        c_sheet.merge_cells('A38:G38')
        c_sheet['I38'] = f'{round(i_0, 2)} %.'
        c_sheet['A39'] = 'pеальный ток холостого хода, A'
        c_sheet.merge_cells('A39:G39')
        c_sheet['A40'] = f'I_0а=(i_0а∙I_1фном)/100=({round(i_0a, 2)}∙{round(i1n)})/100='
        c_sheet.merge_cells('A40:G40')
        c_sheet['I40'] = f'{round(i0a, 5)} A.'
        c_sheet['A41'] = f'I_0а=(i_0p∙I_1фном)/100=({round(i_0p, 2)}∙{round(i1n)})/100='
        c_sheet.merge_cells('A41:G41')
        c_sheet['I41'] = f'{round(i0p, 5)} A.'
        c_sheet['A42'] = f'I_0=(i_0∙I_1фном)/100=({round(i_0, 2)}∙{round(i1n)})/100='
        c_sheet.merge_cells('A42:G42')
        c_sheet['I42'] = f'{round(i0, 5)} A.'
        c_sheet['A43'] = f'cosφ_0={round(i0a, 5)}/{round(i0, 5)} ='
        c_sheet.merge_cells('A43:G43')
        c_sheet['I43'] = f'{round(cosp0, 3)}'
        c_sheet['A44'] = f'sinφ_0={round(i0p, 5)}/{round(i0, 5)} ='
        c_sheet.merge_cells('A44:G44')
        c_sheet['I44'] = f'{round(sinp0, 3)}'

        c_sheet['A46'] = 'б) параметры схемы замещения '
        c_sheet.merge_cells('A46:G46')
        c_sheet['A47'] = f'z_0=U_1фном/I_0 ={round(u1fn, 2)}/{round(i0)}='
        c_sheet.merge_cells('A47:G47')
        c_sheet['I47'] = f'{round(z0)} Ом'
        c_sheet['A48'] = f'r_0=z_0∙cosφ_0={round(z0, 2)}∙{round(cosp0, 3)}='
        c_sheet.merge_cells('A48:G48')
        c_sheet['I48'] = f'{round(r0, 2)} Ом'
        c_sheet['A49'] = f'x_0=√(z_0^2-r_0^2 )=√({round(z0, 2)}^2-{round(r0, 2)}^2 )='
        c_sheet.merge_cells('A49:G49')
        c_sheet['I49'] = f'{round(x0, 2)} Ом'

    return px, qx, i0a, i0p, i0, cosp0, i_0, bc, z0, r0, x0


def step_4(u1fn, i1n, i2n, sp1, sp2, d1, a1, a12, a2, l, w1, w2, s, al, f, r_sheet, c_sheet):
    # Short circuit experience.
    c = 3
    if s <= 100:
        kdob = 1.03
    elif 160 <= s <= 630:
        kdob = 1.06
    else:
        kdob = 1.12
    d_avg1 = round(d1 + 2 * a1, 2)
    d_avg2 = round(d_avg1 + 2 * a12 + a2 / 2, 2)
    j1 = i1n / sp1
    j2 = i2n / sp2
    if al:
        g01 = 8.47 * c * d_avg1 * w1 * sp1 * 10 ** (-5)
        posn1 = 12.75 * g01 * j1 ** 2

        g02 = 8.47 * c * d_avg2 * w2 * sp2 * 10 ** (-5)
        posn2 = 12.75 * g02 * j2 ** 2

        c_sheet['A64'] = f'G_01=8.47∙С∙D_ср1∙w_1∙S_пр1∙10^(-5)=8.47∙3∙{d_avg1}∙{w1}∙{sp1}∙10^(-5)='
        c_sheet.merge_cells('A64:G64')
        c_sheet['I64'] = f'{round(g01, 2)}  кг'
        c_sheet['A65'] = f'G_02=8.47∙С∙D_ср2∙w_2∙S_пр2∙10^(-5)=8.47∙3∙{d_avg2}∙{w2}∙{sp2}∙10^(-5)='
        c_sheet.merge_cells('A65:G65')
        c_sheet['I65'] = f'{round(g02, 2)}  кг'
        c_sheet['A70'] = f'P_осн1=12.75∙J_1^2∙G_01=12.75∙{round(j1, 2)}^2∙{round(g01, 2)}='
        c_sheet.merge_cells('A70:G70')
        c_sheet['I70'] = f'{round(posn1, 2)}  Вт'
        c_sheet['A71'] = f'P_осн2=12.75∙J_2^2∙G_02=12.75∙{round(j2, 2)}^2∙{round(g02, 2)}='
        c_sheet.merge_cells('A71:G71')
        c_sheet['I71'] = f'{round(posn2, 2)}  Вт'
    else:
        g01 = 28 * c * d_avg1 * w1 * sp1 * 10 ** (-5)
        posn1 = 2.4 * g01 * j1 ** 2

        g02 = 28 * c * d_avg2 * w2 * sp2 * 10 ** (-5)
        posn2 = 2.4 * g02 * j2 ** 2

        c_sheet['A64'] = f'G_01=28∙c∙d_avg1∙w1∙sp1∙10^(-5)=28∙3∙{round(d_avg1, 2)}∙{round(w1)}∙{sp1}∙10^(-5)='
        c_sheet.merge_cells('A64:G64')
        c_sheet['I64'] = f'{round(g01, 2)}  кг'
        c_sheet['A65'] = f'G_02=28∙c∙d_avg2∙w2∙sp2∙10^(-5)=28∙3∙{round(d_avg2, 2)}∙{round(w2)}∙{sp2}∙10^(-5)='
        c_sheet.merge_cells('A65:G65')
        c_sheet['I65'] = f'{round(g02, 2)}  кг'
        c_sheet['A70'] = f'P_осн1=2.4∙J_1^2∙G_01=2.4∙{round(j1, 2)}^2∙{round(g01, 2)}='
        c_sheet.merge_cells('A70:G70')
        c_sheet['I70'] = f'{round(posn1, 2)}  Вт'
        c_sheet['A71'] = f'P_осн2=2.4∙J_2^2∙G_02=2.4∙{round(j2, 2)}^2∙{round(g02, 2)}='
        c_sheet.merge_cells('A71:G71')
        c_sheet['I71'] = f'{round(posn2, 2)}  Вт'

    posn = posn1 + posn2
    pk = posn * kdob

    c_sheet['A60'] = 'Плотность тока в обмотках'
    c_sheet.merge_cells('A60:I60')
    c_sheet['A61'] = f'J_1=I_1ном/S_пр1 ={round(i1n, 2)}/{round(sp1, 2)}='
    c_sheet.merge_cells('A61:G61')
    c_sheet['I61'] = f'{round(j1, 2)}  А/мм^2'
    c_sheet['A62'] = f'J_2=I_2ном/S_пр2 ={round(i2n, 2)}/{round(sp2, 2)}='
    c_sheet.merge_cells('A62:G62')
    c_sheet['I62'] = f'{round(j2, 2)}  А/мм^2'
    c_sheet['A63'] = 'Масса обмоток (обмотки алюминиевые)'
    c_sheet.merge_cells('A63:G63')
    c_sheet['A66'] = 'С=3 – число стержней сердечника;'
    c_sheet.merge_cells('A66:G66')
    c_sheet['A67'] = f'D_ср1=D_1+2a_1={d1}+2∙{a1}='
    c_sheet.merge_cells('A67:G67')
    c_sheet['I67'] = f'{round(d_avg1, 2)}  см'
    c_sheet['A68'] = f'D_ср2=D_1+2a_1+2a_12+a_2/2={d1}+2∙{a1}+2∙{a12}+{a2}/2='
    c_sheet.merge_cells('A68:G68')
    c_sheet['I68'] = f'{round(d_avg2, 2)}  см'
    c_sheet['A69'] = 'Потери короткого замыкания рассчитываются для каждой обмотки и затем суммируются:'
    c_sheet.merge_cells('A69:I69')
    c_sheet['A72'] = f'P_осн=P_осн1+P_осн2={round(posn1, 2)}+{round(posn2, 2)}='
    c_sheet.merge_cells('A72:G72')
    c_sheet['I72'] = f'{round(posn, 2)}  Вт'
    c_sheet['A73'] = 'Потери короткого замыкания трансформатора'
    c_sheet.merge_cells('A73:I73')
    c_sheet['A74'] = f'P_к=P_осн∙k_доб={round(posn, 2)}∙{kdob}='
    c_sheet.merge_cells('A74:G74')
    c_sheet['I74'] = f'{round(pk, 2)}  Вт'

    usv = u1fn / w1
    ss = s / c
    d12 = d1 + 2 * a1 + a12
    b = 3.1415 * d12 / l
    ap = a12 + (a1 + a2) / 3
    kp = 0.95
    uap = pk / (10 * s)
    upp = 7.92 * f * ss * b * ap * kp / (usv ** 2 * 10 ** 3)
    ukp_1 = (uap ** 2 + upp ** 2) ** (1 / 2)

    c_sheet['A75'] = f'u_а%={round(pk, 2)}/(10∙{s})='
    c_sheet.merge_cells('A75:G75')
    c_sheet['I75'] = f'{round(uap, 2)}  %'

    c_sheet['A76'] = f'U_в\'-напряжение в одном витке U_в\'=U_1ф/w_1 ={round(u1fn, 2)}/{w1}='
    c_sheet.merge_cells('A76:G76')
    c_sheet['I76'] = f'{round(usv, 2)}'
    c_sheet['A77'] = f'мощность на один стержень: S\'=S/С= '
    c_sheet.merge_cells('A77:G77')
    c_sheet['I77'] = f'{round(ss, 2)} кВА'
    c_sheet['A78'] = f'd_12=D_1+2a_1+a_12={d1}+2∙{a1}+{a12}='
    c_sheet.merge_cells('A78:G78')
    c_sheet['I78'] = f'{round(d12, 2)} см'
    c_sheet['A79'] = f'β=(π∙d_12)/l=(π∙{d12})/{l}= '
    c_sheet.merge_cells('A79:G79')
    c_sheet['I79'] = f'{round(b, 2)}'
    c_sheet['A80'] = f'a_р=a_12+(a_1+a_2)/3={a12}+({a1}+{a2})/3='
    c_sheet.merge_cells('A80:G80')
    c_sheet['I80'] = f'{round(ap, 2)} см'
    c_sheet['A81'] = f'k_р='
    c_sheet.merge_cells('A81:G81')
    c_sheet['I81'] = f'{round(kp, 2)}'

    c_sheet['A82'] = f'u_р%=(7.92∙f∙S\'∙β∙a_р∙k_р∙10^(-3))/(U\'_в^2 )='
    c_sheet.merge_cells('A82:G82')
    c_sheet['A83'] = f'u_р%=(7.92∙50∙{round(ss, 2)}∙{round(b, 2)}∙{round(ap, 2)}∙0.95∙10^(-3))/{round(usv, 2)}^2 ='
    c_sheet.merge_cells('A83:G83')
    c_sheet['I83'] = f'{round(upp, 2)}  %'
    c_sheet['A84'] = f'u_к%=√(u_а%^2+u_р%^2)=√({round(uap, 2)}^2+{round(upp, 2)}^2 )='
    c_sheet.merge_cells('A84:G84')
    c_sheet['I84'] = f'{round(ukp_1, 2)}  %'

    uk = u1fn * ukp_1 / 100
    uka = u1fn * uap / 100
    ukp = u1fn * upp / 100
    c_sheet['A85'] = f'U_к=(U_1фном∙u_к%)/100=({round(u1fn, 2)}∙{round(uk, 2)})/100='
    c_sheet.merge_cells('A85:G85')
    c_sheet['I85'] = f'{round(uk, 2)} В'
    c_sheet['A86'] = f'U_кa=(U_1фном∙u_кa%)/100=({round(u1fn, 2)}∙{round(uka, 2)})/100='
    c_sheet.merge_cells('A86:G86')
    c_sheet['I86'] = f'{round(uka, 2)} В'
    c_sheet['A87'] = f'U_кp=(U_1фном∙u_кp%)/100=({round(u1fn, 2)}∙{round(ukp, 2)})/100='
    c_sheet.merge_cells('A87:G87')
    c_sheet['I87'] = f'{round(ukp, 2)} В'

    zk = uk / i1n
    rk = uka / i1n
    xk = ukp / i1n
    cospk = uka / ukp
    c_sheet['A88'] = f'z_к=U_к/I_1 ={round(uk, 2)}/{round(i1n, 2)}='
    c_sheet.merge_cells('A88:G88')
    c_sheet['I88'] = f'{round(zk, 2)} Ом'
    c_sheet['A89'] = f'r_к=U_кa/I_1 ={round(ukp, 2)}/{round(i1n, 2)}='
    c_sheet.merge_cells('A89:G89')
    c_sheet['I89'] = f'{round(rk, 2)} Ом'
    c_sheet['A90'] = f'x_к=U_кp/I_1 ={round(uk, 2)}/{round(i1n, 2)}='
    c_sheet.merge_cells('A90:G90')
    c_sheet['I90'] = f'{round(xk, 2)} Ом'
    c_sheet['A91'] = f'cosφ_к=U_ка/U_кр ={round(uka, 2)}/{round(ukp, 2)}='
    c_sheet.merge_cells('A91:G91')
    c_sheet['I91'] = f'{round(cospk, 3)}'

    r_sheet['K5'] = 'Результаты вычислений'
    r_sheet['K6'] = f'{round(pk, 2)} W'
    r_sheet['K7'] = f'{round(ukp_1, 2)} %'

    p2_max, delt_u_max = step_4_tabl1(uka, ukp, r_sheet, c_sheet)
    c_sheet['A97'] = 'Величину вторичного напряжения определим по формуле '
    c_sheet.merge_cells('A97:I97')
    c_sheet['A98'] = f'U_2ф=(U_1фном-ΔU)/K=({round(u1fn, 2)}-ΔU)/25'
    c_sheet.merge_cells('A98:I98')
    c_sheet['A99'] = f'ΔU=K_нг(U_ka∙cosφ_2+U_kp∙sinφ_2)=К_нг({round(uka, 2)}∙cosφ_2+{round(ukp, 2)}∙sinφ_2)'
    c_sheet.merge_cells('A99:I99')

    step_4_tabl2(i2n, u1fn, r_sheet, uka, ukp)

    return p2_max, delt_u_max, zk, rk, xk, pk


def step_4_tabl1(uka, ukp, r_sheet, c_sheet):
    r_sheet['A9'] = 'φ_2'
    r_sheet['B9'] = 'ΔU, V'

    c_sheet['A93'] = 'б) Изменение вторичного напряжения ΔU=f(φ2) при номинальном токе (Кнг=1)'
    c_sheet.merge_cells('A93:I93')
    c_sheet['A94'] = f'ΔU=U_ka∙cosφ_2+U_kp∙sinφ_2={round(uka, 2)}∙cosφ_2+{round(ukp, 2)}∙sinφ_2'
    c_sheet.merge_cells('A94:I94')

    delt_u_max = 0
    p2_max = 0
    for ind, p2 in enumerate([-90, -60, -45, -30, 0, 30, 45, 60, 90]):
        r_sheet[f'A{ind + 10}'] = p2
        p2 = math.radians(p2)
        delt_u = uka * math.cos(p2) + ukp * math.sin(p2)
        if delt_u_max < delt_u:
            delt_u_max = delt_u
            p2_max = math.degrees(p2)

        r_sheet[f'B{ind + 10}'] = round(delt_u, 2)
    for p2 in range(round(p2_max) - 10, round(p2_max) + 10):
        p2 = math.radians(p2)
        delt_u = uka * math.cos(p2) + ukp * math.sin(p2)
        if delt_u_max < delt_u:
            delt_u_max = delt_u
            p2_max = math.degrees(p2)
    c_sheet['A95'] = 'Максимальное изменение напряжения при φ_2=φ_к'
    c_sheet.merge_cells('A95:I95')
    c_sheet[
        'A96'] = f'φ_к={round(p2_max, 2)}˚;   ΔU_max={round(uka, 2)}∙cos{round(p2_max, 2)}˚+{round(ukp, 2)}∙sin{round(p2_max, 2)}˚={round(delt_u_max, 2)}, В'
    c_sheet.merge_cells('A96:I96')
    return p2_max, delt_u_max


def step_4_tabl2(i2n, u1fn, r_sheet, uka, ukp):
    r_sheet['D9'] = 'К_нг'
    r_sheet['E9'] = 'I2, A'
    r_sheet.column_dimensions['F'].width = 14
    r_sheet.column_dimensions['G'].width = 16
    r_sheet.column_dimensions['H'].width = 17
    r_sheet.column_dimensions['I'].width = 17
    r_sheet['F9'] = 'ΔU,V\ncosφ_2=1'
    r_sheet['G9'] = 'U2fn,V\ncosφ_2=1'
    r_sheet['H9'] = 'ΔU,V\ncosφ_2=0.7'
    r_sheet['I9'] = 'U2fn,V\ncosφ_2=0.7'
    for ind, kng in enumerate([0, 0.2, 0.4, 0.6, 0.8, 1, 1.2]):
        r_sheet[f'D{ind + 10}'] = kng
        r_sheet[f'E{ind + 10}'] = round(i2n * kng, 2)
        for cos_val in (1, 0.7):
            delt_u = kng * (uka * cos_val + ukp * math.sin(math.acos(cos_val)))
            u2f = (u1fn - delt_u) / 25
            if cos_val == 1:
                letter = ('F', 'G')
            else:
                letter = ('H', 'I')
            r_sheet[f'{letter[0]}{ind + 10}'] = round(delt_u, 2)
            r_sheet[f'{letter[1]}{ind + 10}'] = round(u2f, 2)


def step_5(u1fn, i1n, i2n, p2_max, delt_u_max, rk, xk, r0, x0, c_sheet):
    p2_max = math.radians(p2_max)
    r1 = rk / 2
    x1 = xk / 2
    rm = r0
    xm = x0
    c_sheet['A101'] = f'R_1=R_2\'=R_k/2={round(rk, 2)}/2='
    c_sheet.merge_cells('A101:G101')
    c_sheet['I101'] = f'{round(r1, 2)}Om'
    c_sheet['A102'] = f'x_1=x_2\'=x_k/2={round(xk, 2)}/2='
    c_sheet.merge_cells('A102:G102')
    c_sheet['I102'] = f'{round(x1, 2)}Om'
    c_sheet['A103'] = f'R_m≈R_0='
    c_sheet.merge_cells('A103:G103')
    c_sheet['I103'] = f'{round(rm, 2)}Om'
    c_sheet['A104'] = f'x_m≈x_0='
    c_sheet.merge_cells('A104:G104')
    c_sheet['I104'] = f'{round(xm, 2)}Om'

    c_sheet['A105'] = f'Будем считать I_0=0. Тогда I ̇_1=-I ̇_2^'
    c_sheet.merge_cells('A105:I105')
    i0 = 0

    u2fs = u1fn - delt_u_max
    c_sheet['A106'] = f'U_2ф\'=U_1фном-ΔU={round(u1fn, 2)} - {round(delt_u_max, 2)}='
    c_sheet.merge_cells('A106:G106')
    c_sheet['I106'] = f'{round(u2fs, 2)} V'

    i2s = i1n
    u2s = u2fs * complex(math.cos(p2_max), math.sin(p2_max))
    c_sheet['A107'] = f'Пусть I_1=I_2\'={round(i1n, 2)}'
    c_sheet.merge_cells('A107:G107')
    c_sheet[
        'A108'] = f'Тогда U_2\'={u2fs}∙e^(j{round(math.degrees(p2_max), 2)}°)={round(u2s.real, 2)} + j{round(u2s.imag, 2)}'
    c_sheet.merge_cells('A108:I108')

    e2s = u2s + i2s * complex(r1, x1)
    i1 = - i2s
    u1 = -e2s + i1 * complex(r1, x1)
    c_sheet['A109'] = f'E ̇_2\'=U ̇_2\'+ϳx_2\' I ̇_2\'+R_2\' I ̇_2\'; '
    c_sheet.merge_cells('A109:I109')
    c_sheet['A110'] = f'R_2\' I ̇_2\'={round(r1, 2)}∙{round(i2s, 2)}={round(r1 * i2s, 2)} В;'
    c_sheet.merge_cells('A110:I110')
    c_sheet['A111'] = f'ϳx_2\' I ̇_2\'=ϳ{round(x1, 2)}∙{round(i2s, 2)}=ϳ{round(x1 * i2s, 2)} В'
    c_sheet.merge_cells('A111:i111')
    c_sheet[
        'A112'] = f'E ̇_2\'={round(u2s.real, 2)} + j{round(u2s.imag, 2)}+{round(r1 * i2s, 2)}+ϳ{round(x1 * i2s, 2)}={round(e2s.real, 2)} + j{round(e2s.imag, 2)} В'
    c_sheet.merge_cells('A112:i112')
    c_sheet['A113'] = f'Под углом π⁄2 откладываем Ф_m. '
    c_sheet.merge_cells('A113:i113')
    c_sheet['A114'] = f'I ̇_1=-I ̇_2\'=-{round(i2s, 2)} В; E ̇_1=E ̇_2\''
    c_sheet.merge_cells('A114:i114')
    c_sheet['A115'] = f'U ̇_1=-E ̇_1+ϳx_1 I ̇_1+R_1 I ̇_1=-{round(e2s.real, 2)} - j{round(e2s.imag, 2)}' \
                      f'+{round(r1, 2)}∙-{round(i2s, 2)}+ϳ{round(x1, 2)}∙-{round(i2s, 2)}= {round(u1.real, 2)}  j{round(u1.imag, 2)}'
    c_sheet.merge_cells('A115:i115')


def step_6(s, px, pk, r_sheet, c_sheet):
    # KPD.
    px *= 1e-3
    pk *= 1e-3

    c_sheet['A117'] = f'Зависимость КПД от нагрузки η=f(P_2 ) '
    c_sheet.merge_cells('A117:i117')
    c_sheet['A118'] = f'для cosφ_2=1 и cosφ_2=0,7 '
    c_sheet.merge_cells('A118:i118')
    c_sheet['A119'] = f'η=1-(P+K_нг^2∙P_k)/(K_нг ∙S_ном∙cosφ_2+P+K_нг^2∙P_k)='
    c_sheet.merge_cells('A119:i119')
    c_sheet['A120'] = f'1-({round(px, 5)}+K_нг^2∙{round(pk, 5)})/(K_нг∙{s}∙cosφ_2+{round(px, 5)}+' \
                      f'K_нг^2∙{round(pk, 5)})'
    c_sheet.merge_cells('A120:i120')

    r_sheet['K9'] = 'Snom,kW'
    r_sheet['L9'] = 'η\ncosφ_2=1'
    r_sheet['M9'] = 'η\ncosφ_2=0.7'
    r_sheet.column_dimensions['L'].width = 14
    r_sheet.column_dimensions['M'].width = 16
    for ind, kng in enumerate([0, 0.2, 0.4, 0.6, 0.8, 1, 1.2]):
        r_sheet[f'K{ind + 10}'] = s * kng
        for cos_val in (1, 0.7):
            if cos_val == 1:
                letter = ('L')
            else:
                letter = ('M')
            r_sheet[f'{letter}{ind + 10}'] = round(1 - (px + kng ** 2 * pk) /
                                                   (kng * s * cos_val + px + kng ** 2 * pk), 4)

    kng_max = (px / pk) ** (1 / 2)
    kpd_max = 1 - (px + kng_max ** 2 * pk) / (kng_max * s + px + kng_max * pk)
    c_sheet['A121'] = f'Максимальное значение КПД при cosφ_2=1 '
    c_sheet.merge_cells('A121:i121')
    c_sheet['A122'] = f'при '
    c_sheet.merge_cells('A122:i122')
    c_sheet['A123'] = f'K_нг=√(P_0/P_k )=√({round(px, 5)}/{round(pk, 5)})={round(kng_max, 3)}'
    c_sheet.merge_cells('A123:i123')
    c_sheet['A124'] = f'η_max={round(kpd_max, 4)}'
    c_sheet.merge_cells('A124:i124')


def step_7(i1n, uk, rk, xk, c_sheet):
    # Knocking current.
    ikust = i1n * 100 / uk
    kud = 1 + math.exp(-(math.pi * rk / xk))
    ikm = kud * ikust * 2 ** (1 / 2)

    c_sheet['A126'] = f'Ударный ток короткого замыкания '
    c_sheet.merge_cells('A126:i126')
    c_sheet['A127'] = f'I_к уст=I_1ном∙100/u_k ={round(i1n, 2)}∙100/{round(uk, 2)}={round(ikust, 2)} А ' \
                      f'ток установившегося КЗ '
    c_sheet.merge_cells('A127:i127')
    c_sheet['A128'] = f'k_уд=(1+e^(-(π∙R_k)/(ω∙L_k )) )=(1+e^(-(π∙{round(rk, 2)})/{round(xk, 2)}))={round(kud, 4)}'
    c_sheet.merge_cells('A128:i128')
    c_sheet['A129'] = f'i_км=k_уд∙√2∙I_к уст={round(ikust, 2)}∙√2∙{round(kud, 4)}={round(ikm, 2)} А'
    c_sheet.merge_cells('A129:i129')


def charts(c_sheet, charts_sheet):

    x_data_px = openpyxl.chart.Reference(c_sheet, min_col=2, min_row=2, max_row=8)

    px_u1fn = openpyxl.chart.ScatterChart()
    px_u1fn.title = 'Px_U1fn'
    px_u1fn.style = 13
    px_u1fn.y_axis.title = 'Px'
    px_u1fn.x_axis.title = 'U1fn'
    y_data_px_u1fn = openpyxl.chart.Reference(c_sheet, min_col=3, min_row=2, max_row=9)
    px_u1fn_series = openpyxl.chart.Series(y_data_px_u1fn, x_data_px)
    px_u1fn.series.append(px_u1fn_series)
    line = px_u1fn.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(px_u1fn, 'A1')

    i0_u1fn = openpyxl.chart.ScatterChart()
    i0_u1fn.title = 'i0_U1fn'
    i0_u1fn.style = 13
    i0_u1fn.y_axis.title = 'i0'
    i0_u1fn.x_axis.title = 'U1fn'
    y_data_i0_u1fn = openpyxl.chart.Reference(c_sheet, min_col=7, min_row=2, max_row=8)
    i0_u1fn_series = openpyxl.chart.Series(y_data_i0_u1fn, x_data_px)
    i0_u1fn.series.append(i0_u1fn_series)
    line = i0_u1fn.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(i0_u1fn, 'j1')

    i0a_u1fn = openpyxl.chart.ScatterChart()
    i0a_u1fn.title = 'i0a_U1fn'
    i0a_u1fn.style = 13
    i0a_u1fn.y_axis.title = 'i0a'
    i0a_u1fn.x_axis.title = 'U1fn'
    y_data_i0a_u1fn = openpyxl.chart.Reference(c_sheet, min_col=6, min_row=2, max_row=8)
    i0a_u1fn_series = openpyxl.chart.Series(y_data_i0a_u1fn, x_data_px)
    i0a_u1fn.series.append(i0a_u1fn_series)
    line = i0a_u1fn.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(i0a_u1fn, 'A17')

    i0p_u1fn = openpyxl.chart.ScatterChart()
    i0p_u1fn.title = 'i0p_U1fn'
    i0p_u1fn.style = 13
    i0p_u1fn.y_axis.title = 'i0p'
    i0p_u1fn.x_axis.title = 'U1fn'
    y_data_i0p_u1fn = openpyxl.chart.Reference(c_sheet, min_col=5, min_row=2, max_row=8)
    i0p_u1fn_series = openpyxl.chart.Series(y_data_i0p_u1fn, x_data_px)
    i0p_u1fn.series.append(i0p_u1fn_series)
    line = i0p_u1fn.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(i0p_u1fn, 'j17')

    c0p_u1fn = openpyxl.chart.ScatterChart()
    c0p_u1fn.title = 'c0p_U1fn'
    c0p_u1fn.style = 13
    c0p_u1fn.y_axis.title = 'i0'
    c0p_u1fn.x_axis.title = 'U1fn'
    y_data_c0p_u1fn = openpyxl.chart.Reference(c_sheet, min_col=8, min_row=2, max_row=8)
    c0p_u1fn_series = openpyxl.chart.Series(y_data_c0p_u1fn, x_data_px)
    c0p_u1fn.series.append(c0p_u1fn_series)
    line = c0p_u1fn.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(c0p_u1fn, 'a32')

    sn_n = openpyxl.chart.ScatterChart()
    sn_n.title = 'Snom_KPD'
    sn_n.style = 13
    sn_n.y_axis.title = 'Snom'
    sn_n.x_axis.title = 'KPD'
    sn_n.y_axis.scaling.min = 0.950
    sn_n.y_axis.scaling.max = 0.990
    x_data_sn = openpyxl.chart.Reference(c_sheet, min_col=11, min_row=10, max_row=16)
    for col in range(12, 14):
        y_data_sn = openpyxl.chart.Reference(c_sheet, min_col=col, min_row=10, max_row=16)
        sn_n_series = openpyxl.chart.Series(y_data_sn, x_data_sn)
        sn_n.series.append(sn_n_series)
    line1 = sn_n.series[0]
    line1.marker.symbol = 'dot'
    line1.graphicalProperties.line.width = 20000
    line1.graphicalProperties.line.solidFill = '00bbff'
    line1.marker.graphicalProperties.line.solidFill = '000000'
    line2 = sn_n.series[1]
    line2.marker.symbol = 'dot'
    line2.graphicalProperties.line.width = 20000
    line2.graphicalProperties.line.solidFill = 'ff0000'
    line2.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(sn_n, 'j32')

    du_ph2 = openpyxl.chart.ScatterChart()
    du_ph2.title = 'du_ph2'
    du_ph2.style = 13
    du_ph2.y_axis.title = 'deltU'
    du_ph2.x_axis.title = 'ph2'
    du_ph2.x_axis.scaling.min = -90
    du_ph2.x_axis.scaling.max = 90
    x_data_ph2 = openpyxl.chart.Reference(c_sheet, min_col=1, min_row=10, max_row=18)
    y_data_du = openpyxl.chart.Reference(c_sheet, min_col=2, min_row=10, max_row=18)
    du_ph2_series = openpyxl.chart.Series(y_data_du, x_data_ph2)
    du_ph2.series.append(du_ph2_series)
    line = du_ph2.series[0]
    line.marker.symbol = 'dot'
    line.graphicalProperties.line.width = 20000
    line.graphicalProperties.line.solidFill = 'ff0000'
    line.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(du_ph2, 't1')

    u2fn_i2 = openpyxl.chart.ScatterChart()
    u2fn_i2.title = 'U2fn_I2'
    u2fn_i2.style = 13
    u2fn_i2.y_axis.title = 'U2fn'
    u2fn_i2.x_axis.title = 'I2'
    x_data_i2 = openpyxl.chart.Reference(c_sheet, min_col=5, min_row=10, max_row=16)
    for col in (7, 9):
        y_data_u2fn = openpyxl.chart.Reference(c_sheet, min_col=col, min_row=10, max_row=16)
        u2fn_i2_series = openpyxl.chart.Series(y_data_u2fn, x_data_i2)
        u2fn_i2.series.append(u2fn_i2_series)
    line1 = u2fn_i2.series[0]
    line1.marker.symbol = 'dot'
    line1.graphicalProperties.line.width = 20000
    line1.graphicalProperties.line.solidFill = '00bbff'
    line1.marker.graphicalProperties.line.solidFill = '000000'
    line2 = u2fn_i2.series[1]
    line2.marker.symbol = 'dot'
    line2.graphicalProperties.line.width = 20000
    line2.graphicalProperties.line.solidFill = 'ff0000'
    line2.marker.graphicalProperties.line.solidFill = '000000'
    charts_sheet.add_chart(u2fn_i2, 't17')





def main():
    # get data from excel sheet.
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Лист1']
    sheet_a = wb['Лист2']
    accessory = get_accessory_table(sheet_a)
    kd = 1.25
    f = 50
    for row in range(1, 45):
        rb = openpyxl.Workbook()
        rb.create_sheet()
        rb.create_sheet()
        rb.create_sheet()
        r_sheet = rb.active
        u_sheet = rb['Sheet1']
        c_sheet = rb['Sheet2']
        charts_sheet = rb['Sheet3']

        data = get_data(sheet, row)
        v, s, u1n, u2n, w1, w2, sp1, sp2, d1, a1, a2, a12, l, d, pc, pi, hc, hi, c, pk, px, uk, i0 = data

        if 21 < v <= 31:
            al = False
        else:
            al = True
        data_in_sheet(al, v, s, u1n, u2n, w1, w2, sp1, sp2, d1, a1, a2, a12, l, d, pc, pi,
                      hc, hi, c, pk, px, uk, i0, u_sheet)
        u1fn, u2fn, i1n, i2n, k = step_1(u1n, u2n, s, w1, w2, c_sheet)
        a1_a2 = sorted([a1, a2])
        if u1n > u2n:
            a2, a1 = a1_a2
        else:
            a1, a2 = a1_a2


        step_2([d1, a2, a1, a12, l, d, hc, hi, c], v) # scetch

        r0, x0, px_x = step_3(r_sheet, s, i1n, u1fn, w1, pc, pi, hc, c, d, f, kd, accessory, c_sheet)

        r_sheet['J1'] = 'Контрольные данные'
        r_sheet['J2'] = f'Px = {px}W'
        r_sheet['J3'] = f'i0 = {i0}%'

        r_sheet.column_dimensions['J'].width = 22
        r_sheet.column_dimensions['K'].width = 24
        r_sheet['J5'] = 'Контрольные данные'
        r_sheet['J6'] = f'Pk = {pk}W'
        r_sheet['J7'] = f'uk% = {uk}%'
        p2_max, delt_u_max, zk, rk, xk, pk_k = step_4(u1fn, i1n, i2n, sp1, sp2, d1, a1, a12, a2, l,
                                                w1, w2, s, al, f, r_sheet, c_sheet)

        step_5(u1fn, i1n, i2n, p2_max, delt_u_max, rk, xk, r0, x0, c_sheet)

        step_6(s, px_x, pk_k, r_sheet, c_sheet)

        step_7(i1n, uk, rk, xk, c_sheet)

        charts(r_sheet, charts_sheet)

        rb.save(f'result//var{row}.xlsx')


main()
